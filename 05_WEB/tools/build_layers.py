#!/usr/bin/env python3
"""Build Karben WebGIS GeoJSON layers from local GPKG and KMZ sources."""

from __future__ import annotations

import json
import math
import io
import sqlite3
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
import ezdxf
import shapefile
from pyproj import Transformer
from shapely import wkb
from shapely.geometry import LineString, Point, Polygon, box, mapping, shape
from shapely.ops import transform, unary_union


ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = ROOT.parent
DATA_DIR = ROOT / "05_WEB" / "data"

GPKG_LAYERS = [
    ("../01_Datos_Proyecto/Leitung.gpkg", "leitung", "leitung.geojson"),
    ("../01_Datos_Proyecto/TORRES.gpkg", "tragwerk", "masten.geojson"),
    ("../01_Datos_Proyecto/Seilzugflächen.gpkg", "Seilzugflächen", "arbeitsflaechen.geojson"),
    ("../01_Datos_Proyecto/Arbeitsflaechen_Zusatz.gpkg", "Gerüste", "gerueste.geojson"),
    ("../01_Datos_Proyecto/Arbeitsflaechen_Zusatz.gpkg", "Netz", "netz.geojson"),
    ("../01_Datos_Proyecto/Arbeitsflaechen_Zusatz.gpkg", "Sperrungen", "sperrungen.geojson"),
    ("../01_Datos_Proyecto/Zuwegung_temporär.gpkg", "Zuwegung_temporär", "zuwegung_vorhanden.geojson"),
    ("../01_Datos_Proyecto/Zuwegung_temporär_empty_metadata.gpkg", "Zuwegung_temporär_empty_metadata", "zuwegung_temporaer.geojson"),
    ("../01_Datos_Proyecto/Beschilderung.gpkg", "Beschilderung_Punkte", "beschilderung.geojson"),
    ("../01_Datos_Proyecto/buffer_leitung_800m.gpkg", "buffer_leitung_800m", "buffer_leitung_800m.geojson"),
]

KMZ_LAYERS = [
    ("02_CAD/Leitung.kmz", "leitung_kmz.geojson"),
    ("02_CAD/Tragwerk.kmz", "masten_kmz.geojson"),
]

AUTOCAD_DXF = ("../03_Recursos/Grua Richard.dxf", "autocad_dxf.geojson")

CATASTRO_OVERVIEW_PROPS = {
    "flstnrzae",
    "flstnrnen",
    "flstkennz",
    "gemarkung",
    "flur",
    "gemeinde",
    "kreis",
}

ALKIS_WFS_URL = "https://www.gds.hessen.de/wfs2/aaa-suite/cgi-bin/alkis/vereinf/wfs"
ALKIS_TILE_SIZE_M = 900


def gpkg_wkb(blob: bytes) -> bytes:
    if blob[:2] != b"GP":
        return blob
    flags = blob[3]
    envelope_code = (flags >> 1) & 0b111
    envelope_sizes = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}
    return blob[8 + envelope_sizes.get(envelope_code, 0):]


def transformer_for_srs(srs_id: int):
    if srs_id == 4326:
        return None
    return Transformer.from_crs(f"EPSG:{srs_id}", "EPSG:4326", always_xy=True)


def convert_gpkg(
    gpkg_path: Path,
    table: str,
    output_name: str,
    *,
    attr_filter: set[str] | None = None,
    simplify_tolerance: float | None = None,
    filter_geom=None,
) -> int:
    con = sqlite3.connect(gpkg_path)
    con.row_factory = sqlite3.Row
    geom_row = con.execute(
        "select column_name, srs_id from gpkg_geometry_columns where table_name = ?",
        (table,),
    ).fetchone()
    if not geom_row:
        raise ValueError(f"No geometry column found for {gpkg_path}:{table}")

    geom_col = geom_row["column_name"]
    transformer = transformer_for_srs(int(geom_row["srs_id"]))
    cols = [row["name"] for row in con.execute(f'pragma table_info("{table}")')]
    attr_cols = [col for col in cols if col != geom_col]
    select_cols = ", ".join([f'"{geom_col}"'] + [f'"{col}"' for col in attr_cols])
    rows = con.execute(f'select {select_cols} from "{table}"').fetchall()

    features = []
    for row in rows:
        geom_blob = row[geom_col]
        if not geom_blob:
            continue
        geom = wkb.loads(gpkg_wkb(geom_blob))
        if geom.is_empty:
            continue
        if transformer:
            geom = transform(transformer.transform, geom)
        if filter_geom is not None and not geom.intersects(filter_geom):
            continue
        if simplify_tolerance:
            geom = geom.simplify(simplify_tolerance, preserve_topology=True)
        props = {}
        for col in attr_cols:
            if attr_filter is not None and col not in attr_filter:
                continue
            value = row[col]
            if isinstance(value, bytes):
                value = value.hex()
            props[col] = value
        features.append({"type": "Feature", "properties": props, "geometry": mapping(geom)})

    con.close()
    out = DATA_DIR / output_name
    out.write_text(json.dumps({
        "type": "FeatureCollection",
        "name": Path(output_name).stem,
        "source": str(gpkg_path),
        "features": features,
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return len(features)


def parse_coord_text(text: str):
    coords = []
    for raw in text.split():
        parts = raw.split(",")
        if len(parts) >= 2:
            coords.append([float(parts[0]), float(parts[1])])
    return coords


def properties_from_placemark(pm, ns):
    props = {}
    name = pm.findtext("k:name", default="", namespaces=ns)
    if name:
        props["name"] = name
    description = pm.findtext("k:description", default="", namespaces=ns)
    if description:
        props["description"] = description
    for data in pm.findall(".//k:Data", ns):
        key = data.attrib.get("name")
        value = data.findtext("k:value", default="", namespaces=ns)
        if key:
            props[key] = value
    return props


def kmz_geometry(pm, ns):
    point = pm.find(".//k:Point/k:coordinates", ns)
    if point is not None and point.text:
        coords = parse_coord_text(point.text)
        return Point(coords[0]) if coords else None
    line = pm.find(".//k:LineString/k:coordinates", ns)
    if line is not None and line.text:
        coords = parse_coord_text(line.text)
        return LineString(coords) if len(coords) >= 2 else None
    poly = pm.find(".//k:Polygon//k:outerBoundaryIs/k:LinearRing/k:coordinates", ns)
    if poly is not None and poly.text:
        coords = parse_coord_text(poly.text)
        return Polygon(coords) if len(coords) >= 4 else None
    return None


def convert_kmz(kmz_path: Path, output_name: str) -> int:
    with zipfile.ZipFile(kmz_path) as zf:
        kml_name = next(name for name in zf.namelist() if name.lower().endswith(".kml"))
        root = ET.fromstring(zf.read(kml_name))
    ns = {"k": "http://www.opengis.net/kml/2.2"}
    features = []
    for pm in root.findall(".//k:Placemark", ns):
        geom = kmz_geometry(pm, ns)
        if geom is None or geom.is_empty:
            continue
        features.append({
            "type": "Feature",
            "properties": properties_from_placemark(pm, ns),
            "geometry": mapping(geom),
        })
    out = DATA_DIR / output_name
    out.write_text(json.dumps({
        "type": "FeatureCollection",
        "name": Path(output_name).stem,
        "source": str(kmz_path),
        "features": features,
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return len(features)


def convert_autocad_dxf(dxf_path: Path, output_name: str) -> int:
    doc = ezdxf.readfile(dxf_path)
    transformer = Transformer.from_crs("EPSG:25832", "EPSG:4326", always_xy=True)
    features = []

    def to_wgs84(point):
        lng, lat = transformer.transform(float(point[0]), float(point[1]))
        return [lng, lat]

    def add_feature(geometry_type: str, coordinates, entity, extra=None):
        props = {
            "layer": entity.dxf.layer,
            "entity": entity.dxftype(),
        }
        if extra:
            props.update(extra)
        features.append({
            "type": "Feature",
            "properties": props,
            "geometry": {"type": geometry_type, "coordinates": coordinates},
        })

    def polyline_points(entity):
        if entity.dxftype() == "LWPOLYLINE":
            return [(point[0], point[1]) for point in entity.get_points()]
        return [(vertex.dxf.location.x, vertex.dxf.location.y) for vertex in entity.vertices]

    for entity in doc.modelspace():
        entity_type = entity.dxftype()
        if entity_type == "LINE":
            add_feature("LineString", [to_wgs84(entity.dxf.start), to_wgs84(entity.dxf.end)], entity)
        elif entity_type in ("POLYLINE", "LWPOLYLINE"):
            points = polyline_points(entity)
            if len(points) < 2:
                continue
            coords = [to_wgs84(point) for point in points]
            if entity.is_closed and coords[0] != coords[-1]:
                coords.append(coords[0])
            add_feature("LineString", coords, entity, {"closed": bool(entity.is_closed)})
        elif entity_type == "CIRCLE":
            center = entity.dxf.center
            radius = float(entity.dxf.radius)
            coords = [
                to_wgs84((center.x + radius * math.cos(angle), center.y + radius * math.sin(angle)))
                for angle in [2 * math.pi * index / 48 for index in range(49)]
            ]
            add_feature("LineString", coords, entity, {"radius": radius})
        elif entity_type in ("TEXT", "MTEXT"):
            insert = entity.dxf.insert
            text = entity.dxf.text if entity_type == "TEXT" else entity.text
            add_feature("Point", to_wgs84(insert), entity, {"text": text})

    write_geojson(output_name, features, str(dxf_path))
    return len(features)


def load_geojson(name: str) -> dict:
    return json.loads((DATA_DIR / name).read_text(encoding="utf-8"))


def load_filter_geometry(name: str):
    data = load_geojson(name)
    geoms = [shape(feature["geometry"]) for feature in data.get("features", []) if feature.get("geometry")]
    return unary_union(geoms).buffer(0) if geoms else None


def tile_bounds(bounds: tuple[float, float, float, float], size: float):
    minx, miny, maxx, maxy = bounds
    x = math.floor(minx / size) * size
    while x < maxx:
        y = math.floor(miny / size) * size
        while y < maxy:
            yield (x, y, x + size, y + size)
            y += size
        x += size


def fetch_alkis_tile(bounds_25832: tuple[float, float, float, float]) -> bytes:
    bbox = ",".join(f"{value:.3f}" for value in bounds_25832) + ",EPSG:25832"
    params = {
        "SERVICE": "WFS",
        "VERSION": "2.0.0",
        "REQUEST": "GetFeature",
        "TYPENAMES": "ave:Flurstueck",
        "SRSNAME": "EPSG:25832",
        "BBOX": bbox,
        "OUTPUTFORMAT": "application/x-zip-shapefile",
    }
    url = ALKIS_WFS_URL + "?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=90) as response:
        return response.read()


def reader_from_zip(zip_bytes: bytes) -> shapefile.Reader:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        shp_name = next(name for name in names if name.lower().endswith(".shp"))
        shx_name = next(name for name in names if name.lower().endswith(".shx"))
        dbf_name = next(name for name in names if name.lower().endswith(".dbf"))
        return shapefile.Reader(
            shp=io.BytesIO(zf.read(shp_name)),
            shx=io.BytesIO(zf.read(shx_name)),
            dbf=io.BytesIO(zf.read(dbf_name)),
            encoding="utf-8",
        )


def normalize_props(fields: list[str], record) -> dict:
    props = {}
    for key, value in zip(fields, record):
        clean_key = str(key).lower()
        if isinstance(value, str):
            value = value.strip()
        props[clean_key] = value
    return props


def build_catastro_from_wfs() -> dict[str, int]:
    buffer_wgs84 = load_filter_geometry("buffer_leitung_800m.geojson")
    if buffer_wgs84 is None:
        return {
            "catastro_flurstueck.geojson": write_geojson("catastro_flurstueck.geojson", [], "ALKIS WFS Hessen"),
            "catastro_flurstueck_overview.geojson": write_geojson("catastro_flurstueck_overview.geojson", [], "ALKIS WFS Hessen"),
        }

    to_25832 = Transformer.from_crs("EPSG:4326", "EPSG:25832", always_xy=True)
    to_4326 = Transformer.from_crs("EPSG:25832", "EPSG:4326", always_xy=True)
    buffer_25832 = transform(to_25832.transform, buffer_wgs84)
    features_by_key = {}

    tiles = [bounds for bounds in tile_bounds(buffer_25832.bounds, ALKIS_TILE_SIZE_M) if box(*bounds).intersects(buffer_25832)]
    for index, bounds in enumerate(tiles, 1):
        zip_bytes = fetch_alkis_tile(bounds)
        reader = reader_from_zip(zip_bytes)
        fields = [field[0] for field in reader.fields[1:]]
        for shape_record in reader.iterShapeRecords():
            props = normalize_props(fields, shape_record.record)
            geom_25832 = shape(shape_record.shape.__geo_interface__)
            if geom_25832.is_empty or not geom_25832.intersects(buffer_25832):
                continue
            geom = transform(to_4326.transform, geom_25832)
            if geom.is_empty or not geom.intersects(buffer_wgs84):
                continue
            parcel_key = props.get("flstkennz") or props.get("gml_id") or f"{index}-{len(features_by_key)}"
            features_by_key[str(parcel_key)] = {
                "type": "Feature",
                "properties": props,
                "geometry": mapping(geom),
            }
        print(f"ALKIS WFS tile {index}/{len(tiles)} -> {len(features_by_key)} Flurstuecke")

    features = list(features_by_key.values())
    overview_features = []
    for feature in features:
        props = {key: value for key, value in feature["properties"].items() if key in CATASTRO_OVERVIEW_PROPS}
        geom = shape(feature["geometry"]).simplify(0.000035, preserve_topology=True)
        overview_features.append({"type": "Feature", "properties": props, "geometry": mapping(geom)})

    source = "ALKIS WFS Hessen ave:Flurstueck, tiled by buffer_leitung_800m"
    return {
        "catastro_flurstueck.geojson": write_geojson("catastro_flurstueck.geojson", features, source),
        "catastro_flurstueck_overview.geojson": write_geojson("catastro_flurstueck_overview.geojson", overview_features, source),
    }


def normalize_key(value) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def json_value(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def parcel_key(gemarkung, flur, flurstueck) -> tuple[str, str, str]:
    return (normalize_key(gemarkung), normalize_key(flur), normalize_key(flurstueck))


def kennzeichen_key(value) -> str:
    return normalize_key(value).replace("_", "")


def flurstueck_label(props: dict) -> str:
    zae = props.get("flstnrzae") or props.get("flurstueck")
    nen = props.get("flstnrnen")
    if zae and nen:
        return f"{zae}/{nen}"
    return str(zae or "").strip()


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
        labels = [normalize_key(value) for value in row]
        if "flur" in labels and any(label in labels for label in ("flurstück", "flurstueck", "flurstuck")):
            return list(row)
    return []


def build_status_genehmigung() -> int:
    catastro_path = DATA_DIR / "catastro_flurstueck.geojson"
    if not catastro_path.exists():
        return write_geojson("status_genehmigung.geojson", [], "No catastro available")

    parcel_index = {}
    kennzeichen_index = {}
    catastro = json.loads(catastro_path.read_text(encoding="utf-8"))
    for feature in catastro.get("features", []):
        props = feature.get("properties", {})
        key = parcel_key(props.get("gemarkung"), props.get("flur"), flurstueck_label(props))
        parcel_index.setdefault(key, feature)
        kennzeichen_index.setdefault(kennzeichen_key(props.get("flstkennz")), feature)

    features = []
    for excel_path in sorted((ROOT / "04_PERMITS").glob("*.xls*")):
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header = find_header_row(ws)
            if not header:
                continue
            header_row = next(
                idx for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True), 1)
                if list(row) == header
            )
            cols = {normalize_key(name): pos for pos, name in enumerate(header)}
            gem_col = cols.get("gemarkung")
            flur_col = cols.get("flur")
            flst_col = cols.get("flurstück") or cols.get("flurstueck") or cols.get("flurstuck")
            flstkennz_col = (
                cols.get("flurstückskennzeichen")
                or cols.get("flurstueckskennzeichen")
                or cols.get("flurstuckskennzeichen")
                or cols.get("flstkennz")
            )
            if (flur_col is None or flst_col is None) and flstkennz_col is None:
                continue
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                gemarkung = row[gem_col] if gem_col is not None and gem_col < len(row) else ""
                flur = row[flur_col] if flur_col is not None and flur_col < len(row) else ""
                flurstueck = row[flst_col] if flst_col is not None and flst_col < len(row) else ""
                flstkennz = row[flstkennz_col] if flstkennz_col is not None and flstkennz_col < len(row) else ""
                source_feature = kennzeichen_index.get(kennzeichen_key(flstkennz)) if flstkennz else None
                if not source_feature:
                    source_feature = parcel_index.get(parcel_key(gemarkung, flur, flurstueck))
                if not source_feature:
                    continue
                props = {str(name): json_value(row[pos]) for pos, name in enumerate(header) if name and pos < len(row)}
                props["source_excel"] = excel_path.name
                props["source_sheet"] = ws.title
                features.append({
                    "type": "Feature",
                    "properties": props,
                    "geometry": source_feature["geometry"],
                })
        wb.close()

    return write_geojson("status_genehmigung.geojson", features, "04_PERMITS Excel matched with ALKIS catastro")


def write_geojson(name: str, features: list[dict], source: str) -> int:
    (DATA_DIR / name).write_text(json.dumps({
        "type": "FeatureCollection",
        "name": Path(name).stem,
        "source": source,
        "features": features,
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return len(features)


def build_logistics_layers() -> dict[str, int]:
    mast_data = load_geojson("masten.geojson")
    mast_features = {}
    for feature in mast_data.get("features", []):
        props = feature.get("properties", {})
        mast_name = str(props.get("Name") or props.get("name") or props.get("mast_nr") or "").strip()
        if mast_name:
            mast_features[mast_name.zfill(3)] = feature

    toitoi_features = []
    for mast in ("042", "048", "051"):
        source = mast_features.get(mast)
        if not source:
            continue
        toitoi_features.append({
            "type": "Feature",
            "properties": {
                "name": f"TOI TOI Mast {mast}",
                "mast": mast,
                "typ": "Sanitaer / TOI TOI",
                "projekt": "Karben 042-051",
            },
            "geometry": source["geometry"],
        })

    baulager_features = [{
        "type": "Feature",
        "properties": {
            "name": "Lieferort / Baulager",
            "adresse": "Otto-Hahn-Strasse 35, 63456 Hanau",
            "typ": "Lieferort",
            "projekt": "Karben 042-051",
        },
        "geometry": {
            "type": "Point",
            "coordinates": [8.896721, 50.111146],
        },
    }]

    erdung_specs = {
        "042": ("Arbeitsbereichserdung", "Zone de Trabajo"),
        "044": ("Induktionserdung", "Parallelismus"),
        "047": ("Induktionserdung / Rollenerde", "Kreuzung"),
        "048": ("Arbeitsbereichserdung", "Zone de Trabajo"),
        "050": ("Induktionserdung / Rollenerde", "Kreuzung"),
        "051": ("Arbeitsbereichserdung", "Zone de Trabajo"),
    }
    erdung_features = []
    for mast, (erdung_typ, grund) in erdung_specs.items():
        source = mast_features.get(mast)
        if not source:
            continue
        erdung_features.append({
            "type": "Feature",
            "properties": {
                "name": f"Erdung Mast {mast}",
                "mast": mast,
                "leitung": "LH-11-3024",
                "abschnitt": "042-051",
                "erdung_typ": erdung_typ,
                "grund": grund,
                "hinweis": "Temporaere Erdung nach Datenblatt Nr. 2 je Phase bei Arbeiten am Mast.",
            },
            "geometry": source["geometry"],
        })

    counts = {
        "toitoi.geojson": write_geojson("toitoi.geojson", toitoi_features, "Masten 042, 048, 051"),
        "baulager.geojson": write_geojson("baulager.geojson", baulager_features, "Nominatim geocode / user supplied address"),
        "erdungskonzept.geojson": write_geojson("erdungskonzept.geojson", erdung_features, "Karben Erdungskonzept 042-051"),
    }
    for name in (
        "rettungspunkte.geojson",
        "ausholzung.geojson",
    ):
        counts[name] = write_geojson(name, [], "Placeholder layer - no project data provided yet")
    counts["status_genehmigung.geojson"] = build_status_genehmigung()
    return counts


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    counts = {}
    for source, table, output in GPKG_LAYERS:
        path = (ROOT / source).resolve()
        counts[output] = convert_gpkg(path, table, output)
    counts.update(build_catastro_from_wfs())
    for source, output in KMZ_LAYERS:
        path = (ROOT / source).resolve()
        counts[output] = convert_kmz(path, output)
    dxf_source, dxf_output = AUTOCAD_DXF
    counts[dxf_output] = convert_autocad_dxf((ROOT / dxf_source).resolve(), dxf_output)
    counts.update(build_logistics_layers())
    (DATA_DIR / "layers_manifest.json").write_text(
        json.dumps(counts, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    for name, count in counts.items():
        print(f"{name}\t{count}")


if __name__ == "__main__":
    main()
